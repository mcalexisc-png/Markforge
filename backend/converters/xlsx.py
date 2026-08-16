"""XLSX converter built on openpyxl.

Converts every worksheet in order into Markdown tables, preserving values
(including cached formula results), hyperlinks, merged cells, defined table
ranges, cell comments, images and chart data (charts become tables).
"""

from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from converters.base import BaseConverter, ConversionContext, CorruptFileError
from document_model.blocks import (
    CommentBlock,
    HeadingBlock,
    ImageBlock,
    ParagraphBlock,
    SheetBreakBlock,
    TableBlock,
    TableCell,
    TableRow,
    TextRun,
)
from document_model.document import Document

MAX_ROWS_PER_SHEET = 2000
MAX_COLS_PER_SHEET = 60


class XlsxConverter(BaseConverter):
    format = "xlsx"
    extensions = (".xlsx",)

    def __init__(self, context: ConversionContext):
        super().__init__(context)
        self._sheet_limit_warning: str | None = None

    def convert(self) -> Document:
        path = Path(self.context.source_path)
        try:
            wb = load_workbook(str(path), data_only=True, read_only=False)
            wb_formulas = load_workbook(str(path), data_only=False, read_only=False)
        except Exception as exc:
            raise CorruptFileError("The XLSX file could not be opened or is corrupted.") from exc

        doc_model = self._build_document(
            {
                "title": path.stem,
                "creator": getattr(wb.properties, "creator", None),
                "created": getattr(wb.properties, "created", None),
                "modified": getattr(wb.properties, "modified", None),
            }
        )
        blocks: list = []
        stats = doc_model.stats
        total = len(wb.worksheets)

        for index, ws in enumerate(wb.worksheets, start=1):
            self.context.progress("extract", index, total, f"Extracting sheet {index} of {total}")
            blocks.append(SheetBreakBlock(sheet_number=index, sheet_name=ws.title))
            self._convert_sheet(ws, wb_formulas[ws.title], index, blocks, doc_model)

        stats.sheets = total
        doc_model.blocks = blocks
        doc_model.stats = stats
        return doc_model

    def _convert_sheet(
        self,
        ws: Worksheet,
        ws_formulas: Worksheet,
        sheet_index: int,
        blocks: list,
        doc_model: Document,
    ) -> None:
        stats = doc_model.stats
        blocks.append(HeadingBlock(level=1, content=[TextRun(text=ws.title)]))
        stats.headings += 1

        # Defined tables take priority: they carry real header rows.
        defined_tables = list(ws.tables.values())
        defined_refs: list[str] = []
        for table in defined_tables:
            try:
                range_ref = table.ref
            except Exception:
                continue
            defined_refs.append(range_ref)
            header, rows = self._extract_range(ws, ws_formulas, range_ref, header_row=True)
            if rows:
                blocks.append(TableBlock(rows=rows, has_header=True))
                stats.tables += 1

        used_range = None
        if ws.max_row and ws.max_column:
            used_range = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        if used_range and ws.max_row > 1 and used_range not in defined_refs:
            header, rows = self._extract_range(ws, ws_formulas, used_range, header_row=False)
            if rows:
                blocks.append(TableBlock(rows=rows, has_header=False))
                stats.tables += 1

        if self._sheet_limit_warning:
            doc_model.warnings.append(
                {
                    "code": "sheet_truncated",
                    "message": self._sheet_limit_warning,
                }
            )
            self._sheet_limit_warning = None

        if ws.merged_cells.ranges:
            doc_model.warnings.append(
                {
                    "code": "merged_cells",
                    "message": f"Sheet '{ws.title}' contains merged cells; only the top-left value is kept.",
                    "detail": f"{len(ws.merged_cells.ranges)} merged range(s)",
                }
            )

        if getattr(ws, "_charts", None):
            self._convert_charts(ws, blocks, doc_model)

        self._extract_sheet_comments(ws, blocks, doc_model)
        self._extract_sheet_images(ws, blocks, doc_model)

    def _extract_range(
        self,
        ws: Worksheet,
        ws_formulas: Worksheet,
        range_ref: str,
        *,
        header_row: bool,
    ) -> tuple[bool, list[TableRow]]:
        try:
            cells = ws[range_ref]
        except Exception:
            return False, []
        if not cells:
            return False, []

        row_count = len(cells)
        col_count = max(len(row) for row in cells)
        if row_count > MAX_ROWS_PER_SHEET or col_count > MAX_COLS_PER_SHEET:
            self.warn_sheet_limit(ws.title, row_count, col_count)
            row_count = min(row_count, MAX_ROWS_PER_SHEET)
            col_count = min(col_count, MAX_COLS_PER_SHEET)
            cells = [list(row)[:col_count] for row in cells[:row_count]]

        # Numeric alignment detection per column.
        aligns: list[str | None] = [None] * col_count
        for row in cells:
            for idx, cell in enumerate(row):
                if idx >= col_count:
                    break
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    aligns[idx] = "right" if aligns[idx] != "right" else "right"
                elif cell.value is not None and aligns[idx] == "right":
                    aligns[idx] = None

        rows: list[TableRow] = []
        merged_fill = self._merged_map(ws, range_ref)
        for row_index, row in enumerate(cells):
            out_cells: list[TableCell] = []
            for col_index in range(col_count):
                if col_index >= len(row):
                    out_cells.append(TableCell(content=[]))
                    continue
                cell: Cell = row[col_index]
                value = cell.value
                if value is None:
                    value = merged_fill.get((row_index, col_index))
                if value is None:
                    value = self._formula_fallback(ws_formulas, cell)
                if value is None:
                    value = ""
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                runs = [TextRun(text=str(value))]
                if cell.hyperlink and cell.hyperlink.target:
                    runs[0].href = cell.hyperlink.target
                out_cells.append(TableCell(content=runs, align=aligns[col_index]))
            rows.append(TableRow(cells=out_cells))
        return True, rows

    def _formula_fallback(self, ws_formulas: Worksheet, cell: Cell) -> str | None:
        """Return the raw formula when no cached result exists."""
        raw = ws_formulas[cell.coordinate].value
        if isinstance(raw, str) and raw.startswith("="):
            return raw
        return None

    def _merged_map(self, ws: Worksheet, range_ref: str) -> dict[tuple[int, int], str]:
        """Map merged-cell fill positions to the top-left value."""
        mapping: dict[tuple[int, int], str] = {}
        min_row = ws[range_ref][0][0].row if ws[range_ref] else 1
        min_col = ws[range_ref][0][0].column if ws[range_ref] else 1
        for rng in ws.merged_cells.ranges:
            if rng.min_row > ws.max_row or rng.max_row < min_row:
                continue
            top = ws.cell(row=rng.min_row, column=rng.min_col)
            value = top.value
            if value is None:
                continue
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) != (rng.min_row, rng.min_col):
                        mapping[(r - min_row, c - min_col)] = str(value)
        return mapping

    def warn_sheet_limit(self, title: str, rows: int, cols: int) -> None:
        self._sheet_limit_warning = (
            f"Sheet '{title}' is large ({rows} rows x {cols} cols) and was truncated "
            f"to {MAX_ROWS_PER_SHEET} rows x {MAX_COLS_PER_SHEET} columns."
        )

    def _convert_charts(self, ws: Worksheet, blocks: list, doc_model: Document) -> None:
        """Render each chart's underlying data ranges as a Markdown table."""
        stats = doc_model.stats
        for chart in getattr(ws, "_charts", []) or []:
            try:
                title = self._chart_title(chart)
                series_list: list[tuple[str, list, list]] = []
                sub_charts = getattr(chart, "_charts", None) or [chart]
                for sub in sub_charts:
                    for index, ser in enumerate(getattr(sub, "ser", []) or [], start=1):
                        name = self._chart_text(ser.tx) or f"Series {index}"
                        cat_ref = self._chart_ref(getattr(ser, "cat", None))
                        val_ref = self._chart_ref(getattr(ser, "val", None))
                        cats = self._read_range(ws, cat_ref) if cat_ref else []
                        vals = self._read_range(ws, val_ref) if val_ref else []
                        if not vals:
                            continue
                        series_list.append((name, cats, vals))

                if not series_list:
                    continue
                rows: list[TableRow] = []
                header = [TableCell(content=[TextRun(text="Category")])]
                header.extend(TableCell(content=[TextRun(text=name)]) for name, _, _ in series_list)
                rows.append(TableRow(cells=header))
                row_count = max((len(vals) for _, _, vals in series_list), default=0)
                for row_index in range(row_count):
                    cats = series_list[0][1]
                    category = cats[row_index] if row_index < len(cats) else row_index + 1
                    cells = [TableCell(content=[TextRun(text=str(category))])]
                    for _, _, vals in series_list:
                        value = vals[row_index] if row_index < len(vals) else ""
                        cells.append(TableCell(content=[TextRun(text=self._format_value(value))]))
                    rows.append(TableRow(cells=cells))

                if title:
                    blocks.append(ParagraphBlock(content=[TextRun(text=f"Chart: {title}")]))
                    stats.paragraphs += 1
                blocks.append(TableBlock(rows=rows, has_header=True))
                stats.tables += 1
                stats.charts += 1
            except Exception:
                doc_model.warnings.append(
                    {
                        "code": "chart_extraction_failed",
                        "message": f"A chart in sheet '{ws.title}' could not be parsed and was skipped.",
                        "severity": "warning",
                    }
                )

    @staticmethod
    def _chart_title(chart) -> str:
        title = getattr(chart, "title", None)
        if title is None:
            return ""
        if isinstance(title, str):
            return title.strip()
        try:

            tx = title.tx
            if isinstance(tx, str):
                return tx.strip()
            if tx is not None and getattr(tx, "rich", None) is not None:
                parts = []
                for paragraph in tx.rich.p:
                    for run in getattr(paragraph, "r", []) or []:
                        parts.append(run.t or "")
                text = "".join(parts)
                if text.strip():
                    return text.strip()
        except Exception:
            pass
        return ""

    @staticmethod
    def _chart_text(datasource) -> str | None:
        """Series/category label text from a DataSource (rich or ref)."""
        if datasource is None:
            return None
        try:
            rich = getattr(datasource, "rich", None)
            if rich is not None:
                parts = []
                for paragraph in getattr(rich, "p", []) or []:
                    for run in getattr(paragraph, "r", []) or []:
                        parts.append(run.t or "")
                text = "".join(parts).strip()
                if text:
                    return text
            ref = getattr(datasource, "strRef", None) or getattr(datasource, "numRef", None)
            if ref is not None and getattr(ref, "f", None):
                return None
        except Exception:
            pass
        return None

    @staticmethod
    def _chart_ref(datasource) -> str | None:
        """Cell range address from a chart DataSource, or None."""
        if datasource is None:
            return None
        try:
            for ref in (getattr(datasource, "strRef", None), getattr(datasource, "numRef", None)):
                if ref is not None and getattr(ref, "f", None):
                    return ref.f
        except Exception:
            pass
        return None

    def _read_range(self, ws: Worksheet, address: str) -> list:
        """Read a sheet!$A$1:$A$5 style range into a flat value list."""
        try:
            if "!" in address:
                _, coord = address.rsplit("!", 1)
            else:
                coord = address
            min_col, min_row, max_col, max_row = range_boundaries(coord)
        except Exception:
            return []
        values: list = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                except Exception:
                    continue
                if cell.value is not None:
                    values.append(cell.value)
        return values

    @staticmethod
    def _format_value(value) -> str:
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    def _extract_sheet_comments(self, ws: Worksheet, blocks: list, doc_model: Document) -> None:
        """Append cell comments as CommentBlocks after the sheet tables."""
        stats = doc_model.stats
        comments: list[tuple[str, str, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                comment = getattr(cell, "comment", None)
                if comment is None:
                    continue
                text = getattr(comment, "text", None) or ""
                if not text.strip():
                    continue
                author = getattr(comment, "author", "") or ""
                comments.append((cell.coordinate, author, text.strip()))
        for coordinate, author, text in comments:
            content = [TextRun(text=line.strip()) for line in text.splitlines() if line.strip()]
            if content:
                blocks.append(
                    CommentBlock(
                        content=content,
                        author=f"{author} ({coordinate})" if author else coordinate,
                    )
                )
                stats.comments += 1

    def _extract_sheet_images(self, ws: Worksheet, blocks: list, doc_model: Document) -> None:
        stats = doc_model.stats
        images = getattr(ws, "_images", None) or []
        if not images:
            return
        archive = zipfile.ZipFile(self.context.source_path)
        sheet_rels = self._sheet_rels(ws)
        for image in images:
            try:
                ref = image.ref
            except AttributeError:
                continue
            media = sheet_rels.get(ref)
            data = None
            if media:
                try:
                    data = archive.read(f"xl/{media}")
                except KeyError:
                    data = None
            if data is None:
                try:
                    data = image._data() if hasattr(image, "_data") else None
                except Exception:
                    data = None
            if not data:
                continue
            ext = (image.format or "png").lower() if hasattr(image, "format") else "png"
            rel = self.context.save_image(data, ext)
            if rel:
                stats.images += 1
                blocks.append(ImageBlock(path=rel, alt=""))
        archive.close()

    def _sheet_rels(self, ws: Worksheet) -> dict[str, str]:
        """Map rIds to xl/media paths from the worksheet relationship file."""
        try:
            sheet_path = Path(ws._current_row and ws.title and "xl/worksheets/sheet1.xml")
        except Exception:
            sheet_path = Path("xl/worksheets/sheet1.xml")
        if ws.parent is not None and getattr(ws.parent, "_sheets", None):
            sheet_index = None
            for idx, w in enumerate(ws.parent._sheets, start=1):
                if w is ws:
                    sheet_index = idx
                    break
            if sheet_index is not None:
                sheet_path = Path(f"xl/worksheets/sheet{sheet_index}.xml")
        rels_path = sheet_path.parent / f"_rels/{sheet_path.name}.rels"
        import xml.etree.ElementTree as ET

        out: dict[str, str] = {}
        try:
            with zipfile.ZipFile(self.context.source_path) as archive:
                root = ET.fromstring(archive.read(rels_path.as_posix()))
            ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            for rel in root.findall("r:Relationship", ns):
                rid = rel.get("Id")
                target = rel.get("Target", "")
                if "media" in target:
                    out[rid] = target.lstrip("/")
        except Exception:
            pass
        return out
