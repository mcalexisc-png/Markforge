"""Merge-aware XLSX/XLS converter for the MarkItDown engine.

The stock MarkItDown spreadsheet converters go ``pandas.read_excel`` ->
``DataFrame.to_html()`` -> markdownify, which has three confirmed defects:

- An empty cell renders as the literal string ``NaN``.
- A sheet with no data produces a malformed one-cell table instead of a
  clear "this sheet is empty" note.
- A merged cell's non-top-left members silently lose their content: a
  cell spanning ``A1:B1`` with "Q1" merged into "Q2" comes out as just one
  column, "Q2" gone.

This bypasses pandas entirely and reads the workbook's own cell grid
directly (openpyxl for ``.xlsx``, xlrd for the legacy ``.xls`` format),
backfilling every merged range with its top-left value before rendering --
duplicating a merged value across its span is the standard way to represent
a merge in Markdown, which has no merged-cell syntax of its own.

Known limitation, out of scope here: neither reader converts a date-typed
cell to a readable date string; it is rendered as whatever raw value the
library returns (a serial number for xlrd, since legacy BIFF files store
dates as floats with no format context available without extra work).
"""

from __future__ import annotations

import re

from markitdown._base_converter import DocumentConverterResult
from markitdown.converters import XlsConverter, XlsxConverter


def _format_cell_value(value: object) -> str:
    """Render one cell's value as text, never the literal ``"NaN"``/``"None"``."""
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        # xlrd stores every number as a float (legacy BIFF has no int type),
        # so a whole number like 99 would otherwise show as "99.0".
        return str(int(value))
    return str(value)


def _table_row(cells: list[str]) -> str:
    escaped = [re.sub(r"\s+", " ", cell).replace("|", "\\|").strip() for cell in cells]
    return "| " + " | ".join(escaped) + " |"


def _render_sheet_table(grid: list[list[str]], label: str) -> str:
    """Render one sheet's already-formatted grid as a Markdown table."""
    heading = f"## {label}"
    if not any(cell.strip() for row in grid for cell in row):
        return f"{heading}\n\n_(empty sheet)_"
    header, *body = grid
    lines = [heading, "", _table_row(header), _table_row(["---"] * len(header))]
    lines.extend(_table_row(row) for row in body)
    return "\n".join(lines)


def _grid_from_openpyxl(ws) -> list[list[str]]:
    """Build a formatted string grid, backfilling merged ranges."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    grid: list[list[object]] = [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]
    for merged_range in ws.merged_cells.ranges:
        top_value = grid[merged_range.min_row - 1][merged_range.min_col - 1]
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                grid[r - 1][c - 1] = top_value
    return [[_format_cell_value(v) for v in row] for row in grid]


def _grid_from_xlrd(sheet) -> list[list[str]]:
    """Build a formatted string grid, backfilling merged ranges.

    ``sheet.merged_cells`` entries are ``(row_lo, row_hi, col_lo, col_hi)``,
    0-indexed with a half-open upper bound (row_hi/col_hi are exclusive) --
    only populated when the workbook was opened with ``formatting_info=True``.
    """
    grid: list[list[object]] = [
        [sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)
    ]
    for row_lo, row_hi, col_lo, col_hi in sheet.merged_cells:
        top_value = grid[row_lo][col_lo]
        for r in range(row_lo, row_hi):
            for c in range(col_lo, col_hi):
                grid[r][c] = top_value
    return [[_format_cell_value(v) for v in row] for row in grid]


class MergeAwareXlsxConverter(XlsxConverter):
    """``.xlsx`` converter that renders every sheet as a correct Markdown table."""

    def convert(self, file_stream, stream_info, **kwargs):
        import openpyxl

        workbook = openpyxl.load_workbook(file_stream, data_only=True)
        sections = [
            _render_sheet_table(_grid_from_openpyxl(workbook[name]), name)
            for name in workbook.sheetnames
        ]
        return DocumentConverterResult(markdown="\n\n".join(sections).strip())


class MergeAwareXlsConverter(XlsConverter):
    """Same fix as :class:`MergeAwareXlsxConverter`, for legacy ``.xls``."""

    def convert(self, file_stream, stream_info, **kwargs):
        import xlrd

        # formatting_info=True is required for xlrd to populate
        # sheet.merged_cells -- without it the list is silently empty.
        book = xlrd.open_workbook(
            file_contents=file_stream.read(), formatting_info=True
        )
        sections = [
            _render_sheet_table(
                _grid_from_xlrd(book.sheet_by_index(i)), book.sheet_names()[i]
            )
            for i in range(book.nsheets)
        ]
        return DocumentConverterResult(markdown="\n\n".join(sections).strip())
